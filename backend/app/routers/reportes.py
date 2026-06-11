from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, case
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from app.database import get_db
from app import models, schemas
from app.auth import require_admin

router = APIRouter(prefix="/api/reportes", tags=["Reportes"])


@router.get("/dashboard", response_model=schemas.DashboardStats)
def dashboard_stats(
    fecha_desde: date | None = None,
    fecha_hasta: date | None = None,
    db: Session = Depends(get_db),
    _: tuple = Depends(require_admin),
):
    if not fecha_desde:
        fecha_desde = date.today().replace(day=1)
    if not fecha_hasta:
        fecha_hasta = date.today()

    turnos = db.query(models.Turno).filter(
        models.Turno.fecha >= fecha_desde,
        models.Turno.fecha <= fecha_hasta,
    )
    total = turnos.count()
    completados = turnos.filter(models.Turno.estado == "completado").count()
    cancelados = turnos.filter(models.Turno.estado == "cancelado").count()
    ausentes = turnos.filter(models.Turno.estado == "ausente").count()
    programados = turnos.filter(models.Turno.estado == "programado").count()

    # Ingresos estimados
    ingresos = (
        db.query(func.coalesce(func.sum(models.Tarifa.precio_base), 0))
        .select_from(models.Turno)
        .join(models.Medico, models.Turno.id_medico == models.Medico.id)
        .join(models.Tarifa, models.Tarifa.especialidad == models.Medico.especialidad, isouter=True)
        .filter(
            models.Turno.fecha >= fecha_desde,
            models.Turno.fecha <= fecha_hasta,
            models.Turno.estado == "completado",
        )
        .scalar()
    )

    atendidos = completados + ausentes
    tasa = (completados / atendidos * 100) if atendidos > 0 else 0

    return schemas.DashboardStats(
        total_turnos_mes=total,
        completados=completados,
        cancelados=cancelados,
        ausentes=ausentes,
        programados=programados,
        ingresos_estimados=Decimal(str(ingresos or 0)),
        tasa_asistencia=round(tasa, 1),
    )


@router.get("/por-especialidad", response_model=list[schemas.ReporteEspecialidad])
def reporte_por_especialidad(
    fecha_desde: date | None = None,
    fecha_hasta: date | None = None,
    db: Session = Depends(get_db),
    _: tuple = Depends(require_admin),
):
    if not fecha_desde:
        fecha_desde = date.today().replace(day=1)
    if not fecha_hasta:
        fecha_hasta = date.today()

    results = (
        db.query(
            models.Medico.especialidad,
            func.sum(case((models.Turno.estado == "completado", 1), else_=0)).label("completados"),
            func.sum(case((models.Turno.estado == "ausente", 1), else_=0)).label("ausentes"),
            func.sum(case((models.Turno.estado == "cancelado", 1), else_=0)).label("cancelados"),
            func.coalesce(models.Tarifa.precio_base, 0).label("precio_base"),
        )
        .select_from(models.Turno)
        .join(models.Medico, models.Turno.id_medico == models.Medico.id)
        .join(models.Tarifa, models.Tarifa.especialidad == models.Medico.especialidad, isouter=True)
        .filter(
            models.Turno.fecha >= fecha_desde,
            models.Turno.fecha <= fecha_hasta,
        )
        .group_by(models.Medico.especialidad, models.Tarifa.precio_base)
        .all()
    )

    return [
        schemas.ReporteEspecialidad(
            especialidad=r.especialidad,
            completados=r.completados or 0,
            ausentes=r.ausentes or 0,
            cancelados=r.cancelados or 0,
            precio_base=Decimal(str(r.precio_base)),
            ingresos_estimados=Decimal(str((r.completados or 0) * float(r.precio_base))),
        )
        for r in results
    ]


@router.get("/turnos-por-mes", response_model=list[schemas.TurnoPorMes])
def turnos_por_mes(
    db: Session = Depends(get_db),
    _: tuple = Depends(require_admin),
):
    """Returns appointment counts for the last 6 months."""
    meses_es = {
        1: "Ene", 2: "Feb", 3: "Mar", 4: "Abr", 5: "May", 6: "Jun",
        7: "Jul", 8: "Ago", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dic",
    }
    hoy = date.today()
    resultado = []
    for i in range(5, -1, -1):
        # Go back i months
        mes_actual = hoy.month - i
        anio_actual = hoy.year
        while mes_actual <= 0:
            mes_actual += 12
            anio_actual -= 1
        primer_dia = date(anio_actual, mes_actual, 1)
        if mes_actual == 12:
            ultimo_dia = date(anio_actual + 1, 1, 1) - timedelta(days=1)
        else:
            ultimo_dia = date(anio_actual, mes_actual + 1, 1) - timedelta(days=1)
        total = (
            db.query(func.count(models.Turno.id))
            .filter(
                models.Turno.fecha >= primer_dia,
                models.Turno.fecha <= ultimo_dia,
            )
            .scalar()
        )
        resultado.append(schemas.TurnoPorMes(
            mes=meses_es[mes_actual],
            total=total or 0,
        ))
    return resultado


@router.get("/ingresos-semanales", response_model=list[schemas.IngresoSemanal])
def ingresos_semanales(
    db: Session = Depends(get_db),
    _: tuple = Depends(require_admin),
):
    """Returns estimated revenue for the last 4 weeks."""
    hoy = date.today()
    resultado = []
    for i in range(3, -1, -1):
        inicio_semana = hoy - timedelta(days=hoy.weekday()) - timedelta(weeks=i)
        fin_semana = inicio_semana + timedelta(days=6)
        ingresos = (
            db.query(func.coalesce(func.sum(models.Tarifa.precio_base), 0))
            .select_from(models.Turno)
            .join(models.Medico, models.Turno.id_medico == models.Medico.id)
            .join(models.Tarifa, models.Tarifa.especialidad == models.Medico.especialidad, isouter=True)
            .filter(
                models.Turno.fecha >= inicio_semana,
                models.Turno.fecha <= fin_semana,
                models.Turno.estado == "completado",
            )
            .scalar()
        )
        resultado.append(schemas.IngresoSemanal(
            semana=f"Sem {4 - i}",
            total=Decimal(str(ingresos or 0)),
        ))
    return resultado


@router.get("/exportar-excel")
def exportar_excel(
    fecha_desde: date | None = None,
    fecha_hasta: date | None = None,
    db: Session = Depends(get_db),
    _: tuple = Depends(require_admin),
):
    if not fecha_desde:
        fecha_desde = date.today().replace(day=1)
    if not fecha_hasta:
        fecha_hasta = date.today()

    turnos = (
        db.query(models.Turno)
        .join(models.Paciente, models.Turno.id_paciente == models.Paciente.id)
        .join(models.Medico, models.Turno.id_medico == models.Medico.id)
        .filter(
            models.Turno.fecha >= fecha_desde,
            models.Turno.fecha <= fecha_hasta,
        )
        .order_by(models.Turno.fecha, models.Turno.hora)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Turnos"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="217346", end_color="217346", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Encabezados
    headers = ["Fecha", "Hora", "Paciente", "Médico", "Especialidad", "Estado", "Motivo"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Datos
    for row_idx, turno in enumerate(turnos, 2):
        ws.cell(row=row_idx, column=1, value=turno.fecha.strftime("%d/%m/%Y")).border = thin_border
        ws.cell(row=row_idx, column=2, value=turno.hora.strftime("%H:%M")).border = thin_border
        ws.cell(row=row_idx, column=3, value=turno.paciente.nombre).border = thin_border
        ws.cell(row=row_idx, column=4, value=turno.medico.nombre).border = thin_border
        ws.cell(row=row_idx, column=5, value=turno.medico.especialidad).border = thin_border
        ws.cell(row=row_idx, column=6, value=turno.estado.value if turno.estado else "").border = thin_border
        ws.cell(row=row_idx, column=7, value=turno.motivo or "").border = thin_border

    # Ajustar ancho de columnas
    column_widths = [12, 8, 30, 30, 20, 14, 40]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    # Generar archivo
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"turnos_{fecha_desde.strftime('%Y%m%d')}_{fecha_hasta.strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
